import time
import ddddocr
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoAlertPresentException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import os
import cv2
import numpy as np
import mysql.connector
import base64
import json

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def extract_gatepasses(driver, cmr_lots, agr_map):
    driver.get("https://cgpaddyonline.co.in/millmodify25/RptGatepasslist.aspx")
    try:
        # Wait until the CMR radio button is present
        cmr_radio = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "ctl00_Miller_content1_rdo_type_2"))
        )
        driver.execute_script("arguments[0].click();", cmr_radio)
        
        # Wait until the table is updated/loaded
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "ctl00_Miller_content1_gvReport"))
        )
    except TimeoutException:
        print("Timeout waiting for Gatepass CMR table to load.")
        pass
        
    html = driver.page_source
    soup = BeautifulSoup(html, 'html5lib')
    table = soup.find('table', id='ctl00_Miller_content1_gvReport')
    
    gatepass_items = []
    if table:
        trs = table.find('tbody').find_all('tr') if table.find('tbody') else table.find_all('tr')
        for tr in trs:
            tds = tr.find_all('td')
            if len(tds) >= 8:
                date_val = tds[7].get_text(strip=True)
                mobile_val = tds[6].get_text(strip=True)
                agreement_no = tds[2].get_text(strip=True)
                cmr_center_gp = tds[3].get_text(strip=True)
                lot_match = re.search(r'Lot_No-\s*(\d+)', mobile_val)
                if lot_match:
                    lot_no = lot_match.group(1)
                    passno = ""
                    print_link = tr.find('a', string=re.compile('प्रिंट करे'))
                    if print_link and 'passno=' in print_link.get('href', ''):
                        passno_match = re.search(r'passno=([^&\'"]+)', print_link['href'])
                        if passno_match:
                            passno = passno_match.group(1)
                            
                    gatepass_items.append({
                        "lotNo": lot_no,
                        "date": date_val,
                        "passno": passno,
                        "agreement": agr_map.get(agreement_no, agreement_no),
                        "cmr_center_gp": cmr_center_gp
                    })

    gatepasses = []
    for item in gatepass_items:
        lot_no = item["lotNo"]
        if lot_no in cmr_lots:
            gatepasses.append({
                "lotNo": lot_no,
                "date": item["date"],
                "quantity": cmr_lots[lot_no]["quantity"],
                "approvalDate": cmr_lots[lot_no]["approval_date"],
                "commodity": cmr_lots[lot_no]["commodity"],
                "bagYear": cmr_lots[lot_no]["bag_year"],
                "cmrCenter": cmr_lots[lot_no]["cmr_center"],
                "agreement": item["agreement"],
                "status": "Approved"
            })
        else:
            quantity = 290.0
            commodity = "Unknown"
            if item["passno"]:
                try:
                    driver.get(f"https://cgpaddyonline.co.in/millmodify25/RptGatepassPrint_CMR.aspx?passno={item['passno']}")
                    time.sleep(1.5)
                    print_soup = BeautifulSoup(driver.page_source, 'html5lib')
                    weight_span = print_soup.find('span', id='lbl_weight')
                    if weight_span:
                        quantity = float(re.sub(r'[^\d.]', '', weight_span.get_text(strip=True)))
                    pool_span = print_soup.find('span', id='Label1')
                    if pool_span:
                        commodity = pool_span.get_text(strip=True)
                except:
                    pass
            
            gatepasses.append({
                "lotNo": lot_no,
                "date": item["date"],
                "quantity": quantity,
                "approvalDate": "-",
                "commodity": commodity,
                "bagYear": "-",
                "cmrCenter": item["cmr_center_gp"],
                "agreement": item["agreement"],
                "status": "Pending"
            })
            
    return gatepasses

def solve_captcha(driver):
    ocr = ddddocr.DdddOcr(show_ad=False)
    # Wait for captcha to be visible
    captcha_img = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "ImageCaptcha"))
    )
    captcha_bytes = captcha_img.screenshot_as_png
    
    # Pre-process using OpenCV to improve accuracy
    try:
        nparr = np.frombuffer(captcha_bytes, np.uint8)
        img = cv2.imdecode(nparr, cv2.IMREAD_GRAYSCALE)
        # Thresholding to make it black and white
        _, threshold = cv2.threshold(img, 150, 255, cv2.THRESH_BINARY)
        success, encoded_img = cv2.imencode('.png', threshold)
        if success:
            captcha_bytes = encoded_img.tobytes()
    except Exception as e:
        print(f"OpenCV processing failed, using original image. {e}")
        
    return ocr.classification(captcha_bytes)

def login(driver, user_id, password):
    url = "https://cgpaddyonline.co.in/millmodify25/Default.aspx"
    driver.get(url)
    
    max_retries = 10
    for attempt in range(max_retries):
        try:
            print(f"Login attempt {attempt + 1} for user {user_id}...")
            # Wait for username input field to be present
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "txtUser"))
            )
            
            driver.execute_script(f"document.getElementById('txtUser').value = '{user_id}';")
            driver.execute_script(f"document.getElementById('txtpwd').value = '{password}';")
            
            captcha_text = solve_captcha(driver)
            print(f"Solved Captcha: {captcha_text}")
            
            driver.execute_script(f"document.getElementById('txtVerificationCode').value = '{captcha_text}';")
            driver.execute_script("document.getElementById('btncon').click();")
            
            # Wait for alert or URL change
            try:
                WebDriverWait(driver, 5).until(EC.alert_is_present())
                alert = driver.switch_to.alert
                alert.accept()
            except TimeoutException:
                pass
            
            try:
                btn_ok = WebDriverWait(driver, 3).until(
                    EC.element_to_be_clickable((By.ID, "btnOk"))
                )
                print("Active session found. Clicking OK to deactivate other user...")
                driver.execute_script("arguments[0].click();", btn_ok)
                
                try:
                    WebDriverWait(driver, 3).until(EC.alert_is_present())
                    alert = driver.switch_to.alert
                    alert.accept()
                except TimeoutException:
                    pass
                    
                # Wait for the next page to load
                WebDriverWait(driver, 10).until(
                    EC.url_changes(driver.current_url)
                )
            except TimeoutException:
                pass

            if "Default.aspx" not in driver.current_url and "MillLogin.aspx" not in driver.current_url:
                print(f"Login successful for {user_id}!")
                return True
            else:
                print(f"Login failed on attempt {attempt + 1}. Retrying...")
                driver.get(url)

        except Exception as e:
            print(f"Exception during login: {e}")
            driver.get(url)
            
    return False

def html_table_to_excel(html_table_str, ws, start_row=1, exclude_texts=None, max_rows=None):
    soup = BeautifulSoup(html_table_str, 'html5lib')
    table = soup.find('table')
    if not table: return start_row
    
    fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")
    fill_bg = PatternFill(start_color="428BCA", end_color="428BCA", fill_type="solid") # Darker blue for some headers
    font_white = Font(bold=True, color="FFFFFF")
    font_bold = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    row_idx = start_row
    skip_cells = {}
    
    trs = table.find_all('tr', recursive=False)
    if not trs:
        for child in table.find_all(['tbody', 'thead', 'tfoot'], recursive=False):
            trs.extend(child.find_all('tr', recursive=False))
            
    if max_rows is not None:
        trs = trs[:max_rows]
    
    for tr in trs:
        if exclude_texts:
            row_text = tr.get_text()
            if any(ext in row_text for ext in exclude_texts):
                continue
                
        col_idx = 1
        for td in tr.find_all(['td', 'th'], recursive=False):
            while skip_cells.get((row_idx, col_idx), False):
                col_idx += 1
                
            val = td.get_text(strip=True)
            # convert to numbers where possible so Excel treats them as numbers
            try:
                if '.' in val:
                    val = float(val)
                elif val.isdigit():
                    # If it's a very long number (like Bank Guarantee No), keep it as string!
                    if len(val) > 10:
                        pass
                    else:
                        val = int(val)
            except ValueError:
                pass
                
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = align
            
            # Apply styling for headers
            if td.name == 'th' or 'th' in td.get('class', []):
                # if header is deep blue in original (like State Pool/Central Pool)
                if 'State Pool' in str(val) or 'Central Pool' in str(val) or 'मात्रा' in str(val):
                    cell.fill = fill_bg
                    cell.font = font_white
                else:
                    cell.fill = fill
                    cell.font = font_bold
            
            colspan = int(td.get('colspan', 1))
            rowspan = int(td.get('rowspan', 1))
            
            if colspan > 1 or rowspan > 1:
                ws.merge_cells(start_row=row_idx, start_column=col_idx, end_row=row_idx+rowspan-1, end_column=col_idx+colspan-1)
                
                for r in range(rowspan):
                    for c in range(colspan):
                        if r == 0 and c == 0: continue
                        skip_cells[(row_idx + r, col_idx + c)] = True
                        merged_cell = ws.cell(row=row_idx + r, column=col_idx + c)
                        merged_cell.border = border
                        
            col_idx += colspan
        row_idx += 1
        
    return row_idx

from datetime import datetime

def cleanup_bg_table(ws, start_row, end_row, bank_mapping=None):
    if bank_mapping is None: bank_mapping = {}
    """
    Adjusts serial numbers and recalculates total sum for the Bank Guarantee table 
    after 'पोस्टडेटेड चेक' rows have been filtered out.
    """
    # Hardcoded columns based on fixed portal format
    sno_col, type_col, amt_col, rice_col, paddy_col = 1, 2, 3, 6, 7
    
    current_sno = 1
    sum_amt = 0.0
    sum_rice = 0.0
    sum_rice = 0.0
    sum_paddy = 0.0
    bg_list = []
    
    for r in range(start_row, end_row):
        col1_val = str(ws.cell(row=r, column=sno_col).value or "")
        type_val = str(ws.cell(row=r, column=type_col).value or "")
        
        # Check if it is a data row (Bank Guarantee)
        if type_val.strip() == "बैंक गारंटी":
            ws.cell(row=r, column=sno_col, value=current_sno)
            current_sno += 1
            
            try: sum_amt += float(ws.cell(row=r, column=amt_col).value or 0)
            except: pass
            
            try: sum_rice += float(ws.cell(row=r, column=rice_col).value or 0)
            except: pass
            
            try: sum_paddy += float(ws.cell(row=r, column=paddy_col).value or 0)
            except: pass
            
            try:
                bg_no = str(ws.cell(row=r, column=4).value or "").strip()
                bank_name = bank_mapping.get(bg_no, str(ws.cell(row=r, column=5).value or "").strip())
                amt = float(ws.cell(row=r, column=amt_col).value or 0)
                expiry = str(ws.cell(row=r, column=11).value or "").strip()
                
                if amt > 0 and expiry:
                    days_left = 0
                    try:
                        exp_dt = datetime.strptime(expiry, "%d/%m/%Y")
                        days_left = (exp_dt - datetime.now()).days
                    except: pass
                    
                    bg_list.append({
                        "id": bg_no,
                        "bank": bank_name,
                        "amount": amt,
                        "date": expiry,
                        "daysLeft": days_left
                    })
            except: pass
            
        # Check if it is the Total row
        elif "योग" in col1_val or "योग" in type_val:
            ws.cell(row=r, column=amt_col, value=sum_amt).number_format = '0.00'
            ws.cell(row=r, column=rice_col, value=sum_rice).number_format = '0.00'
            ws.cell(row=r, column=paddy_col, value=sum_paddy).number_format = '0.00'
            ws.cell(row=r, column=paddy_col, value=sum_paddy).number_format = '0.00'
            
    # Sort BGs by days left
    bg_list.sort(key=lambda x: x['daysLeft'])
            
    return sum_amt, bg_list

def extract_rice_balance(ws, last_data_row):
    """
    Extracts the 'Total Balance of Rice to deposit' by summing the last 8 columns 
    of the 'योग' row in the Rice Data table.
    """
    total_row = last_data_row - 1
    for r in range(last_data_row, max(1, last_data_row - 10), -1):
        if "योग" in str(ws.cell(row=r, column=1).value or ""):
            total_row = r
            break
            
    target = sum(float(ws.cell(row=total_row, column=c).value or 0) for c in range(12, 20) if ws.cell(row=total_row, column=c).value)
    deposited = sum(float(ws.cell(row=total_row, column=c).value or 0) for c in range(20, 28) if ws.cell(row=total_row, column=c).value)
    
    mapping = {
        28: "State Pool (Fortified - Common)",
        29: "State Pool (Fortified - Grade A)",
        30: "State Pool (Normal - Common)",
        31: "State Pool (Normal - Grade A)",
        32: "Central Pool (Fortified - Common)",
        33: "Central Pool (Fortified - Grade A)",
        34: "Central Pool (Normal - Common)",
        35: "Central Pool (Normal - Grade A)",
    }
    
    balance = 0.0
    for c in range(28, 36):
        try:
            val = float(ws.cell(row=total_row, column=c).value or 0)
            balance += val
        except:
            pass
            
    agreement_balances = []
    for r in range(total_row - 1, 0, -1):
        c1_val = str(ws.cell(row=r, column=1).value or "").strip()
        if c1_val.isdigit():
            agr_no = str(ws.cell(row=r, column=2).value or "").strip()
            agr_type = str(ws.cell(row=r, column=3).value or "").strip()
            
            agr_details = {
                "agreement_no": agr_no,
                "agreement_type": agr_type,
                "total_pending": 0.0,
                "qualities": {}
            }
            
            for c in range(28, 36):
                try:
                    val = float(ws.cell(row=r, column=c).value or 0)
                    if val > 0:
                        agr_details["qualities"][mapping[c]] = val
                        agr_details["total_pending"] += val
                except:
                    pass
                    
            if agr_details["total_pending"] > 0:
                agreement_balances.append(agr_details)
        else:
            break
            
    agreement_balances.reverse()
            
    return target, deposited, balance, agreement_balances

def cleanup_do_table(ws, start_row, end_row, agr_no=""):
    """
    The portal has broken HTML for the DO table total row (missing empty cells),
    causing the totals to shift left by 1 column.
    This recalculates the mathematical sum agreement-wise and places it in the correct columns.
    """
    total_row_idx = end_row - 1
    
    # Ensure there are enough rows to have headers + total
    if total_row_idx <= start_row + 1:
        return 0.0, 0.0, []
        
    sums = {c: 0.0 for c in range(6, 18)}
    
    pending_dos = []
    
    for r in range(start_row + 2, total_row_idx):
        for c in range(6, 18):
            try:
                val = float(ws.cell(row=r, column=c).value or 0)
                sums[c] += val
            except:
                pass
                
        # Extract pending DO info
        try:
            mota_bal = float(ws.cell(row=r, column=15).value or 0)
            patla_bal = float(ws.cell(row=r, column=16).value or 0)
            total_bal = float(ws.cell(row=r, column=17).value or 0)
            
            if total_bal > 0:
                center = str(ws.cell(row=r, column=2).value or "Unknown Center").strip()
                date = str(ws.cell(row=r, column=5).value or "Unknown Date").strip()
                
                mota_do_qty = float(ws.cell(row=r, column=6).value or 0)
                sarna_do_qty = float(ws.cell(row=r, column=8).value or 0)
                
                if mota_bal > 0:
                    ptype = "Mota/Sarna"
                    if mota_do_qty > 0 and sarna_do_qty == 0:
                        ptype = "Mota"
                    elif sarna_do_qty > 0 and mota_do_qty == 0:
                        ptype = "Sarna"
                    pending_dos.append({
                        "agreement": agr_no,
                        "center": center,
                        "date": date,
                        "type": ptype,
                        "qty": mota_bal
                    })
                if patla_bal > 0:
                    pending_dos.append({
                        "agreement": agr_no,
                        "center": center,
                        "date": date,
                        "type": "Patla",
                        "qty": patla_bal
                    })
        except:
            pass
                
    # Clear the existing shifted total row values
    for c in range(1, 18):
        ws.cell(row=total_row_idx, column=c).value = ""
        
    # Write the new correct totals
    ws.cell(row=total_row_idx, column=2, value="योग :").font = Font(bold=True)
    
    for c in range(6, 18):
        cell = ws.cell(row=total_row_idx, column=c, value=sums[c])
        cell.number_format = '0.00'
        cell.font = Font(bold=True)
        
    return sums[9], sums[17], pending_dos

def scrape_data():
    options = webdriver.ChromeOptions()
    # options.add_argument('--headless')
    driver = webdriver.Chrome(options=options)
    
    # Fetch accounts from MySQL dynamically
    accounts = []
    try:
        print("Connecting to MySQL Database 'cmrs_pro' to fetch active accounts...")
        db_conn = mysql.connector.connect(
            host="gateway01.ap-southeast-1.prod.alicloud.tidbcloud.com",
            port=4000,
            user="47DktuAxv5uhMxU.root",
            password="IgkKsq49O0py7Cxs",
            database="cmrs_pro",
            ssl_verify_cert=False,
            ssl_verify_identity=False
        )
        cursor = db_conn.cursor()
        cursor.execute("SELECT miller_id, portal_password FROM miller_credentials")
        rows = cursor.fetchall()
        for row in rows:
            accounts.append((row[0], row[1]))
        cursor.close()
        db_conn.close()
        print(f"Successfully fetched {len(accounts)} accounts from database: {[a[0] for a in accounts]}")
    except Exception as db_err:
        print(f"Error fetching accounts from database: {db_err}")
        # Fallback to hardcoded accounts if database fetch fails
        accounts = [
            ("MA446578", "Raju#14436"),
            ("MA441707", "Raju#14436")
        ]
    
    data_dir = os.path.join(BASE_DIR, "data")
    if not os.path.exists(data_dir):
        os.makedirs(data_dir)
    
    data_found = False
    financial_data = []

    try:
        batch_size = 10
        for idx, (user_id, password) in enumerate(accounts):
            if idx > 0 and idx % batch_size == 0:
                print(f"Batch limit reached ({batch_size} mills processed). Taking a 3-minute break to avoid server overload...")
                time.sleep(180)
                
            miller_rice_target = 0.0
            miller_rice_deposited = 0.0
            miller_rice_balance = 0.0
            rice_qualities = []
            miller_bg_amount = 0.0
            miller_total_do = 0.0
            miller_balance_paddy = 0.0
            miller_pending_dos = []
            cmr_lots = {}
            miller_gatepasses = []
            miller_name_full = user_id
            
            wb = Workbook()
            ws_rice = wb.active
            ws_rice.title = "Rice Data"
            ws_bg = wb.create_sheet("Bank Guarantee")
            ws_do = wb.create_sheet("DO Data")
            ws_cmr = wb.create_sheet("CMR Data")
            ws_fin = wb.create_sheet("Financial Report")
            
            rice_row = 1
            bg_row = 1
            do_row = 1
            cmr_row = 1
            miller_name_full = user_id
            
            wb = Workbook()
            ws_rice = wb.active
            ws_rice.title = "Rice Data"
            ws_bg = wb.create_sheet("Bank Guarantee")
            ws_do = wb.create_sheet("DO Data")
            ws_cmr = wb.create_sheet("CMR Data")
            ws_fin = wb.create_sheet("Financial Report")
            
            rice_row = 1
            bg_row = 1
            do_row = 1
            cmr_row = 1

            
            print(f"\n--- Processing Account {user_id} ---")
            if login(driver, user_id, password):
                data_found = True
                print("Successfully logged in.")
                time.sleep(3)
                
                try:
                    edit_btn = driver.find_element(By.ID, "lbEdit")
                    if edit_btn.is_displayed():
                        print("Found Edit button. Clicking it...")
                        driver.execute_script("arguments[0].click();", edit_btn)
                        time.sleep(5)
                except Exception:
                    print("No Edit button found. Proceeding...")

                print("Navigating to Rice table page...")
                target_url = "https://cgpaddyonline.co.in/millmodify25/RptSubmitableRiceByMiller.aspx"
                driver.get(target_url)
                time.sleep(5)
                
                print("Extracting Rice table data natively...")
                html_rice = driver.page_source
                soup_rice = BeautifulSoup(html_rice, 'html5lib')
                tables_rice = soup_rice.find_all('table')
                
                if tables_rice:
                    target_table_rice = max(tables_rice, key=lambda t: len(t.text))
                    ws_rice.cell(row=rice_row, column=1, value=f"मिलर ID: {user_id} - जमा चावल की मात्रा").font = Font(bold=True, size=14, color="2A5DAB")
                    rice_row += 2
                    rice_row = html_table_to_excel(str(target_table_rice), ws_rice, start_row=rice_row)
                    miller_rice_target, miller_rice_deposited, miller_rice_balance, rice_qualities = extract_rice_balance(ws_rice, rice_row)
                    rice_row += 4
                else:
                    print(f"No Rice data found for {user_id}.")

                print("Extracting Detailed Bank Names from SaveMiller page...")
                driver.get("https://cgpaddyonline.co.in/millmodify25/TypeSecuritySaveMiller.aspx")
                time.sleep(4)
                html_save = driver.page_source
                soup_save = BeautifulSoup(html_save, 'html5lib')
                bank_mapping = {}
                for tr in soup_save.find_all('tr'):
                    tds = tr.find_all('td')
                    if len(tds) >= 6:
                        bg_no = tds[2].get_text(strip=True)
                        bank_name_full = tds[4].get_text(strip=True)
                        if bg_no and bank_name_full:
                            bank_mapping[bg_no] = bank_name_full

                print("Navigating to Bank Guarantee (BG) table page...")
                bg_url = "https://cgpaddyonline.co.in/millmodify25/TypeSecurityCompleteReport.aspx"
                driver.get(bg_url)
                time.sleep(5)
                
                print("Extracting BG table data natively...")
                html_bg = driver.page_source
                soup_bg = BeautifulSoup(html_bg, 'html5lib')
                lblmsg = soup_bg.find('span', id='ctl00_lblmsg')
                if lblmsg:
                    miller_name_full = lblmsg.get_text(strip=True)
                lblmsg = soup_bg.find('span', id='ctl00_lblmsg')
                if lblmsg:
                    miller_name_full = lblmsg.get_text(strip=True)

                tables_bg = soup_bg.find_all('table')
                
                if tables_bg:
                    target_table_bg = max(tables_bg, key=lambda t: len(t.text))
                    ws_bg.cell(row=bg_row, column=1, value=f"मिलर ID: {user_id} - प्रतिभूति (Bank Guarantee)").font = Font(bold=True, size=14, color="2A5DAB")
                    bg_row += 2
                    
                    bg_start = bg_row
                    bg_row = html_table_to_excel(str(target_table_bg), ws_bg, start_row=bg_row, exclude_texts=["पोस्टडेटेड चेक"])
                    
                    miller_bg_amount, miller_bg_list = cleanup_bg_table(ws_bg, bg_start, bg_row, bank_mapping)
                    
                    bg_row += 4
                else:
                    print(f"No BG data found for {user_id}.")

                print("Navigating to DO table page...")
                do_url = "https://cgpaddyonline.co.in/millmodify25/AgreementReconciliationAll.aspx"
                driver.get(do_url)
                time.sleep(5)
                
                try:
                    from selenium.webdriver.support.ui import Select
                    dropdown_elem = driver.find_element(By.ID, "ctl00_Miller_content1_DDAgreementNo")
                    dropdown = Select(dropdown_elem)
                    opts = [{"value": opt.get_attribute("value"), "text": opt.text.strip()} for opt in dropdown.options if opt.get_attribute("value") != "0"]
                    
                    if not opts:
                        print(f"No DO agreements found for {user_id}.")
                    else:
                        print(f"Found {len(opts)} DO agreements. Extracting...")
                        for opt_data in opts:
                            opt_val = opt_data["value"]
                            # Must re-find element after page refresh
                            dropdown = Select(driver.find_element(By.ID, "ctl00_Miller_content1_DDAgreementNo"))
                            dropdown.select_by_value(opt_val)
                            driver.find_element(By.ID, "ctl00_Miller_content1_btnshow").click()
                            time.sleep(5)
                            
                            html_do = driver.page_source
                            soup_do = BeautifulSoup(html_do, 'html5lib')
                            tables_do = soup_do.find_all('table')
                            
                            if len(tables_do) >= 2:
                                ws_do.cell(row=do_row, column=1, value=f"मिलर ID: {user_id} | एग्रीमेंट: {opt_val.strip()}").font = Font(bold=True, size=14, color="2A5DAB")
                                do_row += 2
                                
                                # Extract summary table
                                do_row = html_table_to_excel(str(tables_do[0]), ws_do, start_row=do_row, max_rows=5)
                                do_row += 2
                                
                                # Extract actual DO table
                                do_start = do_row
                                do_row = html_table_to_excel(str(tables_do[1]), ws_do, start_row=do_row)
                                
                                # Clean up and recalculate totals due to broken portal HTML
                                do_tot, do_bal, pending_dos = cleanup_do_table(ws_do, do_start, do_row, opt_val.strip())
                                miller_total_do += do_tot
                                miller_balance_paddy += do_bal
                                miller_pending_dos.extend(pending_dos)
                                
                                do_row += 4
                                
                                # Extract CMR details table
                                target_cmr_table = None
                                for t in tables_do:
                                    if 'अभिस्विकृति' in t.text and 'लॉट क्रमांक' in t.text and not t.find('table'):
                                        target_cmr_table = t
                                        break
                                        
                                if target_cmr_table:
                                    ws_cmr.cell(row=cmr_row, column=1, value=f"मिलर ID: {user_id} | एग्रीमेंट: {opt_val.strip()}").font = Font(bold=True, size=14, color="2A5DAB")
                                    cmr_row += 2
                                    cmr_row = html_table_to_excel(str(target_cmr_table), ws_cmr, start_row=cmr_row)
                                    cmr_row += 4
                                    
                                    # Extract CMR Lots to match against Gate Pass
                                    trs = target_cmr_table.find_all('tr')
                                    for tr in trs:
                                        tds = tr.find_all(['td', 'th'])
                                        if len(tds) >= 10:
                                            lot_text = tds[7].get_text(strip=True)
                                            qty_text = tds[4].get_text(strip=True)
                                            app_date = tds[2].get_text(strip=True)
                                            commodity = tds[8].get_text(strip=True)
                                            type_match = re.search(r'\((.*?)\)', lot_text)
                                            if type_match:
                                                commodity = f"{commodity} ({type_match.group(1)})"
                                            
                                            bag_year = tds[9].get_text(strip=True)
                                            cmr_center = tds[3].get_text(strip=True)
                                            lot_match = re.search(r'(\d+)', lot_text)
                                            if lot_match:
                                                try:
                                                    cmr_lots[lot_match.group(1)] = {
                                                        "quantity": float(qty_text),
                                                        "approval_date": app_date,
                                                        "commodity": commodity,
                                                        "bag_year": bag_year,
                                                        "cmr_center": cmr_center
                                                    }
                                                except: pass
                                    
                except Exception as e:
                    print(f"Failed to process DO/CMR for {user_id}: {e}")
                
                print("Navigating to Gatepass list page...")
                
                agr_map = { a["agreement_no"]: f"{a['agreement_no']} ({a['agreement_type']})" for a in rice_qualities }
                
                try:
                    miller_gatepasses = extract_gatepasses(driver, cmr_lots, agr_map)
                except Exception as e:
                    print(f"Failed to extract gatepasses for {user_id}: {e}")
                    miller_gatepasses = []
                
                # Perform Calculations for Financial Report
                paddy_qty = (miller_rice_balance / 0.676777) + miller_balance_paddy
                paddy_amt = paddy_qty * 2500
                free_bg = miller_bg_amount - paddy_amt
                financial_data.append({
                    'miller_id': user_id,
                    'miller_name_full': miller_name_full,
                    'rice_target': miller_rice_target,
                    'rice_deposited': miller_rice_deposited,
                    'rice_balance': miller_rice_balance,
                    'rice_qualities': rice_qualities,
                    'total_do': miller_total_do,
                    'balance_paddy': miller_balance_paddy,
                    'paddy_qty': paddy_qty,
                    'paddy_amt': paddy_amt,
                    'balance_bg': miller_bg_amount,
                    'free_bg': free_bg,
                    'gatepasses': miller_gatepasses,
                    'pendingDOs': miller_pending_dos,
                    'bgs': miller_bg_list if 'miller_bg_list' in locals() else []
                })
                
                # Write Financial Report for THIS miller
                headers1 = ["मिलर ID", "Total Balance of Rice to deposit", "Paddy Qty", "Paddy Amt.", "Balance BG", "Free Bg"]
                headers2 = ["", "Total Balance of Rice to deposit", "Total Rice Balance to Deposit/67.676777% + Balance Paddy to lift", "Paddy Qty*2500", "Total of BG Amount - Release", "Balance BG - Paddy Amount"]
                for col_num, (h1, h2) in enumerate(zip(headers1, headers2), 1):
                    c1 = ws_fin.cell(row=1, column=col_num, value=h1)
                    c2 = ws_fin.cell(row=2, column=col_num, value=h2)
                    c1.font = Font(bold=True)
                    c1.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    c1.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    c2.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                
                ws_fin.cell(row=3, column=1, value=user_id)
                ws_fin.cell(row=3, column=2, value=miller_rice_balance).number_format = '0.000'
                ws_fin.cell(row=3, column=3, value=paddy_qty).number_format = '0.00'
                ws_fin.cell(row=3, column=4, value=paddy_amt).number_format = '0.00'
                ws_fin.cell(row=3, column=5, value=miller_bg_amount).number_format = '0.00'
                ws_fin.cell(row=3, column=6, value=free_bg).number_format = '0.00'
                
                for ws in [ws_rice, ws_bg, ws_do, ws_cmr, ws_fin]:
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 18
                        
                # Create Folder
                safe_name = re.sub(r'[\\\\/*?:"<>|]', "", miller_name_full)
                # Save localized data to Excel
                miller_dir = os.path.join(data_dir, f"{safe_name} ({user_id})")
                if not os.path.exists(miller_dir):
                    os.makedirs(miller_dir)
                    
                excel_path = os.path.join(miller_dir, f"{user_id}_Report.xlsx")
                try:
                    wb.save(excel_path)
                    print(f"Saved {user_id} Excel to {excel_path}")
                except PermissionError:
                    print(f"Warning: Could not save Excel {excel_path} because the file is open in another program. Dashboard data will still be updated.")
                
                # Create isolated Dashboard data for this miller
                miller_js_data = {
                    user_id: {
                        "name": f"{miller_name_full} ({user_id})",
                        "targetRice": miller_rice_target,
                        "depositedRice": miller_rice_deposited,
                        "riceBalance": miller_rice_balance,
                        "riceQualities": rice_qualities,
                        "totalAllottedPaddy": miller_total_do,
                        "balancePaddy": miller_balance_paddy,
                        "paddyAmt": paddy_amt,
                        "bgAmount": miller_bg_amount,
                        "freeBg": free_bg,
                        "gatePassStatus": miller_gatepasses,
                        "nearestBgs": miller_bg_list if 'miller_bg_list' in locals() else [],
                        "bags": { "new": 0, "old": 0, "pds": 0 }
                    }
                }
                
                # Inject the dashboard app so the folder has a STANDALONE SINGLE HTML FILE!
                index_path = os.path.join(BASE_DIR, "index.html")
                style_path = os.path.join(BASE_DIR, "style.css")
                app_path = os.path.join(BASE_DIR, "app.js")
                leaderboard_path = os.path.join(BASE_DIR, "raipur_leaderboard.json")
                logo_path = os.path.join(BASE_DIR, "logo.png")
                
                if os.path.exists(index_path):
                    with open(index_path, "r", encoding="utf-8") as f_app:
                        app_content = f_app.read()
                    if os.path.exists(style_path):
                        with open(style_path, "r", encoding="utf-8") as f_css:
                            css_str = f_css.read()
                        app_content = app_content.replace('<link rel="stylesheet" href="style.css">', f"<style>\n{css_str}\n</style>")
                        
                    with open(app_path, "r", encoding="utf-8") as f_js:
                        js_app_str = f_js.read()
                        
                    # Load leaderboard
                    leaderboard_str = ""
                    if os.path.exists(leaderboard_path):
                        with open(leaderboard_path, "r", encoding="utf-8") as fl:
                            leaderboard_str = "\nwindow.leaderboardData = " + fl.read() + ";"
                    
                    if os.path.exists(logo_path):
                        with open(logo_path, "rb") as img_f:
                            b64_logo = base64.b64encode(img_f.read()).decode('utf-8')
                            app_content = app_content.replace('src="logo.png"', f'src="data:image/png;base64,{b64_logo}"')
                            
                    js_string = "window.dashboardData = " + json.dumps(miller_js_data) + ";"
                    js_combined = js_string + leaderboard_str + "\n" + js_app_str
                    app_content = app_content.replace('<script src="app.js"></script>', f"<script>\n{js_combined}\n</script>")
                    app_content = app_content.replace('<script src="data.js"></script>', "")
                    
                    out_html = os.path.join(out_dir, f"{safe_name}_Dashboard.html")
                    with open(out_html, "w", encoding="utf-8") as f_out:
                        f_out.write(app_content)
                    print(f"Created standalone dashboard for {user_id} at {out_html}")

                
                driver.delete_all_cookies()
            else:
                 print(f"Could not login for {user_id}")

        if data_found:
            # Generate Global data.js for the UI Dashboard
            import json
            js_data = {
                "combined": {
                    "name": "Combined Operations Overview",
                    "targetRice": 0.0,
                    "depositedRice": 0.0,
                    "riceBalance": 0.0,
                    "riceQualities": [],
                    "totalAllottedPaddy": 0.0,
                    "balancePaddy": 0.0,
                    "paddyAmt": 0.0,
                    "bgAmount": 0.0,
                    "freeBg": 0.0,
                    "gatePassStatus": [],
                    "pendingDOs": [],
                    "nearestBgs": [],
                    "bags": { "new": 0, "old": 0, "pds": 0 }
                }
            }
            
            for fdata in financial_data:
                mid = fdata['miller_id']
                name_prefix = fdata['miller_name_full']
                js_data[mid] = {
                    "name": f"{name_prefix} ({mid})",
                    "targetRice": fdata['rice_target'],
                    "depositedRice": fdata['rice_deposited'],
                    "riceBalance": fdata['rice_balance'],
                    "riceQualities": fdata['rice_qualities'],
                    "totalAllottedPaddy": fdata['total_do'],
                    "balancePaddy": fdata['balance_paddy'],
                    "paddyAmt": fdata['paddy_amt'],
                    "bgAmount": fdata['balance_bg'],
                    "freeBg": fdata['free_bg'],
                    "gatePassStatus": fdata['gatepasses'],
                    "pendingDOs": fdata['pendingDOs'],
                    "nearestBgs": fdata.get('bgs', []),
                    "bags": { "new": 0, "old": 0, "pds": 0 }
                }
                
                # Accumulate for combined
                js_data["combined"]["targetRice"] += fdata['rice_target']
                js_data["combined"]["depositedRice"] += fdata['rice_deposited']
                js_data["combined"]["riceBalance"] += fdata['rice_balance']
                
                js_data["combined"]["riceQualities"].extend(fdata['rice_qualities'])
                js_data["combined"]["totalAllottedPaddy"] += fdata['total_do']
                js_data["combined"]["balancePaddy"] += fdata['balance_paddy']
                js_data["combined"]["paddyAmt"] += fdata['paddy_amt']
                js_data["combined"]["bgAmount"] += fdata['balance_bg']
                js_data["combined"]["freeBg"] += fdata['free_bg']
                js_data["combined"]["gatePassStatus"].extend(fdata['gatepasses'])
                js_data["combined"]["pendingDOs"].extend(fdata['pendingDOs'])
                js_data["combined"]["nearestBgs"].extend(fdata.get('bgs', []))
                
            js_data["combined"]["nearestBgs"].sort(key=lambda x: x.get('daysLeft', 0))
            with open("data.js", "w", encoding="utf-8") as f:
                js_str = "window.dashboardData = " + json.dumps(js_data, indent=4) + ";"
                f.write(js_str + "\n")
                
            # --- MySQL Database Insertion ---
            try:
                print("Connecting to MySQL Database 'cmrs_pro'...")
                db_conn = mysql.connector.connect(
                    host="gateway01.ap-southeast-1.prod.alicloud.tidbcloud.com",
                    port=4000,
                    user="47DktuAxv5uhMxU.root",
                    password="IgkKsq49O0py7Cxs",
                    database="cmrs_pro",
                    ssl_verify_cert=False,
                    ssl_verify_identity=False
                )
                cursor = db_conn.cursor()
                
                for fdata in financial_data:
                    mid = fdata['miller_id']
                    name = fdata['miller_name_full']
                    t_rice = fdata['rice_target']
                    d_rice = fdata['rice_deposited']
                    r_bal = fdata['rice_balance']
                    t_paddy = fdata['total_do']
                    b_paddy = fdata['balance_paddy']
                    p_amt = fdata['paddy_amt']
                    bg_amt = fdata['balance_bg']
                    f_bg = fdata['free_bg']
                    
                    # Insert Miller
                    cursor.execute("""
                        INSERT INTO millers (id, name, target_rice, deposited_rice, rice_balance, total_allotted_paddy, balance_paddy, paddy_amt, bg_amount, free_bg) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE 
                        target_rice=VALUES(target_rice), deposited_rice=VALUES(deposited_rice), rice_balance=VALUES(rice_balance), 
                        total_allotted_paddy=VALUES(total_allotted_paddy), balance_paddy=VALUES(balance_paddy), paddy_amt=VALUES(paddy_amt), 
                        bg_amount=VALUES(bg_amount), free_bg=VALUES(free_bg)
                    """, (mid, name, t_rice, d_rice, r_bal, t_paddy, b_paddy, p_amt, bg_amt, f_bg))
                    
                    # Rice Qualities
                    cursor.execute("DELETE FROM rice_qualities WHERE miller_id = %s", (mid,))
                    for agr in fdata['rice_qualities']:
                        agr_no = agr.get('agreement_no', '')
                        agr_type = agr.get('agreement_type', '')
                        for q_type, qty in agr.get('qualities', {}).items():
                            cursor.execute("INSERT INTO rice_qualities (miller_id, agreement_no, agreement_type, quality_type, quantity) VALUES (%s, %s, %s, %s, %s)", (mid, agr_no, agr_type, q_type, qty))
                            
                    # Gatepasses
                    cursor.execute("DELETE FROM gatepasses WHERE miller_id = %s", (mid,))
                    for gp in fdata['gatepasses']:
                        cursor.execute("INSERT INTO gatepasses (miller_id, lot_no, pass_date, quantity, status, agreement_no, cmr_center, commodity, bag_year, approval_date) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                                       (mid, gp.get('lotNo'), gp.get('date'), gp.get('quantity'), gp.get('status'), gp.get('agreement'), gp.get('cmrCenter'), gp.get('commodity'), gp.get('bagYear'), gp.get('approvalDate')))
                                       
                    # Pending DOs
                    cursor.execute("DELETE FROM pending_dos WHERE miller_id = %s", (mid,))
                    for do in fdata['pendingDOs']:
                        cursor.execute("INSERT INTO pending_dos (miller_id, agreement_no, center, do_date, paddy_type, quantity) VALUES (%s, %s, %s, %s, %s, %s)",
                                       (mid, do.get('agreement'), do.get('center'), do.get('date'), do.get('type'), do.get('qty')))
                                       
                    # Bank Guarantees
                    cursor.execute("DELETE FROM bank_guarantees WHERE miller_id = %s", (mid,))
                    for bg in fdata.get('bgs', []):
                        cursor.execute("INSERT INTO bank_guarantees (miller_id, bank_name, bg_no, amount, valid_date, days_left) VALUES (%s, %s, %s, %s, %s, %s)",
                                       (mid, bg.get('bank'), bg.get('id'), bg.get('amount'), bg.get('date'), bg.get('daysLeft')))
                                       
                    # Update sync time in credentials table
                    cursor.execute("UPDATE miller_credentials SET last_sync_time = NOW() WHERE miller_id = %s", (mid,))
                                       
                db_conn.commit()
                cursor.close()
                db_conn.close()
                print("Successfully updated data in MySQL Database cmrs_pro!")
            except Exception as db_err:
                print(f"Database error: {db_err}")
            # --- End MySQL Insertion ---
            
            # Load leaderboard
            leaderboard_str = ""
            if os.path.exists("raipur_leaderboard.json"):
                with open("raipur_leaderboard.json", "r", encoding="utf-8") as fl:
                    leaderboard_str = "\nwindow.leaderboardData = " + fl.read() + ";"

            # Also create a self-contained Combined Dashboard!
            index_path = os.path.join(BASE_DIR, "index.html")
            style_path = os.path.join(BASE_DIR, "style.css")
            app_path = os.path.join(BASE_DIR, "app.js")
            leaderboard_path = os.path.join(BASE_DIR, "raipur_leaderboard.json")
            logo_path = os.path.join(BASE_DIR, "logo.png")
            combined_path = os.path.join(BASE_DIR, "Combined_Dashboard.html")
            
            if os.path.exists(index_path):
                with open(index_path, "r", encoding="utf-8") as f_app:
                    app_content = f_app.read()
                
                if os.path.exists(style_path):
                    with open(style_path, "r", encoding="utf-8") as f_css:
                        css_str = f_css.read()
                    app_content = app_content.replace('<link href="style.css" rel="stylesheet"/>', f"<style>\n{css_str}\n</style>")
                    app_content = app_content.replace('<link rel="stylesheet" href="style.css">', f"<style>\n{css_str}\n</style>")
                    
                with open(app_path, "r", encoding="utf-8") as f_js:
                    js_app_str = f_js.read()
                
                if os.path.exists(logo_path):
                    with open(logo_path, "rb") as img_f:
                        b64_logo = base64.b64encode(img_f.read()).decode('utf-8')
                        app_content = app_content.replace('src="logo.png"', f'src="data:image/png;base64,{b64_logo}"')
                
                leaderboard_str = ""
                if os.path.exists(leaderboard_path):
                    with open(leaderboard_path, "r", encoding="utf-8") as fl:
                        leaderboard_str = "\nwindow.leaderboardData = " + fl.read() + ";"
                
                js_combined = js_str + leaderboard_str + "\n" + js_app_str
                app_content = app_content.replace('<script src="app.js"></script>', f"<script>\n{js_combined}\n</script>")
                app_content = app_content.replace('<script src="data.js"></script>', "")
                
                with open(combined_path, "w", encoding="utf-8") as f_out:
                    f_out.write(app_content)
                print("Successfully generated Combined_Dashboard.html for sharing!")
                
            print("Successfully exported GLOBAL dashboard UI data!")
            
        else:
            print("\nNo data was collected.")



    except Exception as e:
        print(f"An error occurred: {e}")
    finally:
        driver.quit()

if __name__ == "__main__":
    scrape_data()
